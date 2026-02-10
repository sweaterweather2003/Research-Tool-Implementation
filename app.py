import streamlit as st
from pypdf import PdfReader
import openai
import openpyxl
from io import BytesIO
import re

openai.api_key = "sk-your-openai-api-key-her"

# Standard line items with synonyms
standard_items = {
    "Revenue": ["Net Sales", "Total Revenue", "Sales", "Operating Revenue"],
    "Cost of Goods Sold": ["Cost of Sales", "COGS", "Cost of Revenue"],
    "Gross Profit": ["Gross Margin", "Gross Income"],
    "Operating Expenses": ["Operating Costs", "Selling, General and Administrative Expenses", "SG&A", "Operating Expenses"],
    "Operating Income": ["EBIT", "Operating Profit", "Income from Operations"],
    "Net Income": ["Net Profit", "Profit After Tax", "Bottom Line"],
    # Add more as needed, e.g., "Interest Expense", "Taxes"
}

def extract_text_from_pdf(file):
    reader = PdfReader(file)
    text = ""
    for page in reader.pages:
        text += page.extract_text() + "\n"
    return text

def detect_currency_and_units(text):
    # Pattern matching for currency/units
    currency_match = re.search(r'\b(USD|EUR|INR|GBP)\b', text, re.IGNORECASE)
    units_match = re.search(r'\b(in thousands|in millions|in billions|000s)\b', text, re.IGNORECASE)
    currency = currency_match.group(1).upper() if currency_match else "Unknown"
    units = units_match.group(1) if units_match else "Units not specified"
    return currency, units

def extract_with_llm(text):
    prompt = f"""
    Extract income statement line items from the following text. Map to these standard items: {list(standard_items.keys())}.
    Use synonyms like {standard_items} to match variations.
    Output only what's present; do not hallucinate or infer missing data.
    Handle multiple years if present (e.g., identify years like 2023, 2022).
    Extract exact numeric values (handle commas, negatives in parentheses).
    If ambiguous, note in 'notes'.
    Detect currency/units from text if mentioned.

    Output as JSON:
    {{
        "currency": "str",
        "units": "str",
        "years": ["year1", "year2", ...],
        "line_items": {{
            "Standard Item": {{"values": [val1, val2, ...], "notes": "str"}}
        }}
    }}

    Text: {text[:20000]}  # Truncate if too long; adjust as needed
    """
    
    response = openai.chat.completions.create(
        model="gpt-4o-mini",
        messages=[{"role": "system", "content": "You are a financial extractor."}, {"role": "user", "content": prompt}],
        response_format={"type": "json_object"}
    )
    return response.choices[0].message.content

def create_excel(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Income Statement"

    # Header
    json_data = eval(data)  # Safe since from LLM JSON
    currency, units = json_data["currency"], json_data["units"]
    years = json_data["years"]
    ws.cell(row=1, column=1).value = f"Currency: {currency}, Units: {units}"

    # Table headers
    ws.cell(row=3, column=1).value = "Line Item"
    for col, year in enumerate(years, start=2):
        ws.cell(row=3, column=col).value = year
    ws.cell(row=3, column=len(years)+2).value = "Notes"

    # Data rows
    row_num = 4
    for item, details in json_data["line_items"].items():
        ws.cell(row=row_num, column=1).value = item
        values = details["values"]
        notes = details["notes"]
        for col, val in enumerate(values, start=2):
            if val == "N/A":
                ws.cell(row=row_num, column=col).value = "N/A"
                ws.cell(row=row_num, column=col).fill = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red for missing
            elif "ambiguous" in notes.lower():
                ws.cell(row=row_num, column=col).value = val
                ws.cell(row=row_num, column=col).fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow for ambiguous
            else:
                ws.cell(row=row_num, column=col).value = val
        ws.cell(row=row_num, column=len(years)+2).value = notes
        row_num += 1

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# Streamlit App
st.title("Research Portal: Financial Statement Extractor")

uploaded_files = st.file_uploader("Upload Documents (PDFs)", type="pdf", accept_multiple_files=True)

if uploaded_files:
    st.success(f"Uploaded {len(uploaded_files)} documents.")
    selected_file = st.selectbox("Select document to process:", [f.name for f in uploaded_files])
    
    if st.button("Extract Financial Statements"):
        file = next(f for f in uploaded_files if f.name == selected_file)
        text = extract_text_from_pdf(file)
        
        if text:
            with st.spinner("Extracting..."):
                llm_output = extract_with_llm(text)
                excel_file = create_excel(llm_output)
                st.download_button(
                    label="Download Excel",
                    data=excel_file,
                    file_name="income_statement.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:

            st.error("No text extracted from PDF.")
